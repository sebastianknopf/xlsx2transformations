import json
import logging
import openpyxl as xl

class TransformationGenerator:

    def __init__(self, xl_filename):
        self.xlwb = xl.load_workbook(xl_filename, data_only=True)

        if not 'transformations' in self.xlwb:
            raise ValueError(f"workbook does not contain 'transformations' worksheet")

        self._transformation_sequence = list()
        for transformation in self.xlwb['transformations'].iter_rows(min_row=2, values_only=True):
            self._transformation_sequence.append((
                transformation[0],
                transformation[1]
            ))

        self._transformations = list()

    def generate(self, txt_filename):
        for worksheet in self.xlwb.worksheets:
            if worksheet.title.endswith('.txt'):
                self._add_transformations(worksheet)

        with open(txt_filename, 'w', encoding='utf-8') as txt_file:
            for transformation_type in self._transformation_sequence:
                for transformation in self._transformations:
                    if transformation['op'] == 'add':
                        compare_type = ('add', transformation['obj']['file'])
                    else:
                        compare_type = (transformation['op'], transformation['match']['file'])

                    if set(compare_type) == set(transformation_type):
                        txt_file.write(json.dumps(transformation, ensure_ascii=False))
                        txt_file.write('\n')

    def _add_transformations(self, worksheet):
        headers = list()
        for cell in worksheet.iter_cols(min_row=1, max_row=1, values_only=True):
            if cell[0] != 'op':
                headers.append(str(cell[0]))

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            operation = row[0].strip()

            transformation = dict()
            transformation['op'] = operation

            if operation == 'add':
                obj = dict()
                obj['file'] = worksheet.title

                for index, header in enumerate(headers):
                    obj[header] = self._sanitize_value(row[index + 1])

                transformation['obj'] = obj
            elif operation == 'update':
                update = dict()
                for index, header in enumerate(headers):
                    update[header] = self._sanitize_value(row[index + 1])

                match = dict()
                match['file'] = worksheet.title
                match = match | self._create_file_match(worksheet.title, headers, row)
                
                transformation['update'] = update
                transformation['match'] = match

            elif operation == 'remove':
                match = dict()
                match['file'] = worksheet.title
                match = match | self._create_file_match(worksheet.title, headers, row)

                transformation['match'] = match

            elif operation == 'retain':
                match = dict()
                match['file'] = worksheet.title
                match = match | self._create_file_match(worksheet.title, headers, row)

                transformation['match'] = match

            self._transformations.append(transformation)

    def _create_file_match(self, filename, headers, row):
        if filename == 'agency.txt':
            return self._create_match(['agency_id'], headers, row)
        elif filename == 'routes.txt':
            return self._create_match(['route_id'], headers, row)
        elif filename == 'stops.txt':
            return self._create_file_match(['stop_id'], headers, row)
        elif filename == 'trips.txt':
            return self._create_file_match(['trip_id'], headers, row)
    
    def _create_match(self, keys, headers, row):
        match = dict()
        for key in keys:
            if key in headers:
                match[key] = str(row[headers.index(key) + 1])
        
        return match

    def _sanitize_value(self, value):
        if value is not None:
            return str(value).strip()
        else:
            return ''
